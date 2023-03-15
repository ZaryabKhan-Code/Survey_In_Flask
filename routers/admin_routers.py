from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from flask_bcrypt import check_password_hash
from flask import *
from config.config import *
from utils.admin_utils import *
from models.user_models import *
import pdfkit


admin_router = Blueprint('admin_model', __name__,static_folder='static', template_folder='views')
login_manager2.login_view = 'admin_model.admin_login'


@login_manager2.user_loader
def load_user(user_id):
    return LoadUserById(user_id)


@admin_router.route('/register', methods=['GET', 'POST'])
@login_required
def admin_register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']

        if not username or not email or not password or not confirm_password:
            error_message = 'Please fill out all fields.'
            return render_template('register.html', error_message=error_message)

        if password != confirm_password:
            error_message = 'Passwords do not match.'
            return render_template('register.html', error_message=error_message)

        existing_user = get_admin_username(username)
        if existing_user:
            error_message = 'That username is already taken.'
            return render_template('register.html', error_message=error_message)

        existing_email = get_admin_email(email)
        if existing_email:
            error_message = 'That email is already taken.'
            return render_template('register.html', error_message=error_message)

        new_user = Admin(username=username, email=email, password=password, is_confirmed=False)
        db.session.add(new_user)
        db.session.commit()
        send_verification_email(email, username)
        message = "Registration successful, activation link sent to email"
        return render_template('register.html', message=message)
    return render_template('register.html')

    
def create_default_user():
    email = 'admin@example.com'
    password = 'password123'
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    if not Admin.query.filter_by(email=email).first():
        default_user = Admin(
            username='admin',
            email=email,
            password=hashed_password,
            is_confirmed=True
        )
        db.session.add(default_user)
        db.session.commit()

@admin_router.route('/admin', methods=['POST', 'GET'])
def admin_login():
    success_message = request.args.get('success_message')
    if current_user.is_authenticated:
        return redirect(url_for('admin_model.admin_dashboard'))
    try:
        if request.method == 'POST':
            email = request.form['email']
            password = request.form['pass']
            user = get_admin_email(email)
            if user and bcrypt.check_password_hash(user.password, password):
                login_user(user)
                return redirect(url_for('admin_model.admin_dashboard'))
            if not user.is_confirmed:
                error_message = 'Your account has not been confirmed yet.'
                return render_template('login.html', error_message=error_message, success_message=success_message)
            else:
                error_message = 'Invalid email or password'
                return render_template('login.html', error_message=error_message)
        return render_template('login.html', success_message=success_message)
    except Exception as e:
        error_message = f'Login access denied {e}'
        return render_template('login.html', error_message=error_message, success_message=success_message)

@admin_router.route('/check_email/<email>')
@login_required
def check_email(email):
    email_admin = get_admin_email(email)
    if email_admin:
        return jsonify({'error': 'Email already exists'})
    else:
        return jsonify({})


@admin_router.route('/check_username/<username>')
@login_required
def check_username(username):
    username_admin = get_admin_username(username)
    if username_admin:
        return jsonify({'error': 'Username already exists'})
    else:
        return jsonify({})


@admin_router.route('/activate')
def activate():
    if current_user.is_authenticated:
        return redirect(url_for('admin_model.admin_dashboard'))
    token = request.args.get('token')
    try:
        email = verify_activation_token(token)
        if email is not None:
            activate = get_admin_email(email)
            activate.is_confirmed = True
            db.session.commit()
            delete_token(token)
            flash('Your email has been verified. You can now log in.')
            return render_template('verification_result.html', success=True)
        else:
            return render_template('verification_result.html', success=False)
    except Exception as e:
        return f'Error Creating Token Contact Support'


@admin_router.route('/diplomas/<int:user_id>/<filename>', methods=['GET'])
@login_required
def download_diploma(user_id, filename):
    diploma = Diploma.query.filter_by(user_id=user_id).first()
    if diploma is None:
        abort(404, description='Diploma not found!')

    if filename == diploma.filename_diploma_image:
        file_data = diploma.diploma_image
    elif filename == diploma.filename_identity_proof:
        file_data = diploma.identity_proof
    elif filename == diploma.filename_personal_photo:
        file_data = diploma.personal_photo
    else:
        abort(404, description='File not found!')

    response = make_response(file_data)
    response.headers.set('Content-Type', 'application/octet-stream')
    response.headers.set('Content-Disposition', 'attachment',
                         filename=f'User_Id:{user_id}_'+filename)
    return response


@admin_router.route('/user/show/information')
@login_required
def admin_dashboard():
    try:
        user = User.query.all()
        form = Form.query.first()
        return render_template('dashboard.html', user=user,form=form)
    except Exception as e:
        message = f'Error'
        return render_template('dashboard.html', message=message)


@admin_router.route('/logout')
@login_required
def admin_logout():
    success_message = 'User Successfully Logout'
    logout_user()
    return redirect(url_for('admin_model.admin_login', success_message=success_message))


@admin_router.route('/user/details/<id>', methods=['POST', 'GET'])
@login_required
def user_information(id):
    user = User.query.filter_by(id=id).first()
    other_info = OtherInformation.query.filter_by(user_id=id).first()
    disability = Disability.query.filter_by(user_id=id).first()
    university = University.query.filter_by(user_id=id).first()
    degreeprogram = DegreeProgram.query.filter_by(user_id=id).first()
    vocationaltrainingcenters = Vocationaltrainingcenters.query.filter_by(
        user_id=id).first()
    institution = Institution.query.filter_by(user_id=id).first()
    technicalTraining = TechnicalTraining.query.filter_by(user_id=id).first()
    diploma = Diploma.query.filter_by(user_id=id).first()
    return render_template('user_deatils.html', user_id=id, user=user, other_info=other_info, disability=disability, university=university, institution=institution, degreeprogram=degreeprogram, vocationaltrainingcenters=vocationaltrainingcenters, technicalTraining=technicalTraining, diploma=diploma)


path_wkhtmltopdf = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)




@admin_router.route('/diplomas/<int:user_id>/download/images')
@login_required
def download_zip_images(user_id):
    diploma = Diploma.query.filter_by(user_id=user_id).first()
    if diploma is None:
        abort(404, description='Diploma not found!')

    user = User.query.filter_by(id=user_id).first()
    files = [
        (f'{user.first_name}_diploma_File_' +
         diploma.filename_diploma_image, diploma.diploma_image),
        (f'{user.first_name}_identity_proof_' +
         diploma.filename_identity_proof, diploma.identity_proof),
        (f'{user.first_name}_personal_photo_' +
         diploma.filename_personal_photo, diploma.personal_photo)
    ]

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for filename, data in files:
            zip_file.writestr(filename, data)

    zip_buffer.seek(0)
    response = make_response(zip_buffer.getvalue())
    response.headers.set('Content-Type', 'application/zip')
    response.headers.set('Content-Disposition', 'attachment',
                         filename=f'UserRecord_{user.id_card}_Images_Record.zip')
    return response

@admin_router.route('/diplomas/<int:user_id>/download/PDF')
@login_required
def download_zip_pdf(user_id):
    diploma2 = Diploma.query.filter_by(user_id=user_id).first()

    if diploma2 is None:
        abort(404, description='Diploma not found!')
    user = User.query.filter_by(id=user_id).first()
    other_info = OtherInformation.query.filter_by(user_id=user_id).first()
    disability = Disability.query.filter_by(user_id=user_id).first()
    university = University.query.filter_by(user_id=user_id).first()
    degreeprogram = DegreeProgram.query.filter_by(user_id=user_id).first()
    vocationaltrainingcenters = Vocationaltrainingcenters.query.filter_by(
        user_id=user_id).first()
    institution = Institution.query.filter_by(user_id=user_id).first()
    technicalTraining = TechnicalTraining.query.filter_by(
        user_id=user_id).first()
    diploma = Diploma.query.filter_by(user_id=user_id).first()
    options = {
        'page-size': 'A4',
        'margin-top': '0',
        'margin-right': '0',
        'margin-bottom': '0',
        'margin-left': '0',
    }
    html = render_template('pdfgenerator.html', user_id=id, user=user, other_info=other_info, disability=disability, university=university, institution=institution,
                           degreeprogram=degreeprogram, vocationaltrainingcenters=vocationaltrainingcenters, technicalTraining=technicalTraining, diploma=diploma)
    pdf = pdfkit.from_string(
        html, False, options=options, configuration=config)

    files = [
        (f'{user.first_name}_diploma_File_' +
         diploma2.filename_diploma_image, diploma2.diploma_image),
        (f'{user.first_name}_identity_proof_' +
         diploma2.filename_identity_proof, diploma2.identity_proof),
        (f'{user.first_name}_personal_photo_' +
         diploma2.filename_personal_photo, diploma2.personal_photo),
        (f'{user.first_name}_' + 'Information.pdf', pdf)
    ]

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for filename, data in files:
            zip_file.writestr(filename, data)

    zip_buffer.seek(0)
    response = make_response(zip_buffer.getvalue())
    response.headers.set('Content-Type', 'application/zip')
    response.headers.set('Content-Disposition', 'attachment',
                         filename=f'UserRecord_{user.id_card}_PDF_Record.zip')
    return response


@admin_router.route('/diplomas/<int:user_id>/download/EXCEL', methods=['POST', 'GET'])
@login_required
def download_zip_excel(user_id):
        user = User.query.filter_by(id=user_id).first()
        info = OtherInformation.query.filter_by(user_id=user_id).first()
        dis = Disability.query.filter_by(user_id=user_id).first()
        uni = University.query.filter_by(user_id=user_id).first()
        deg = DegreeProgram.query.filter_by(user_id=user_id).first()
        voc = Vocationaltrainingcenters.query.filter_by(
            user_id=user_id).first()
        inst = Institution.query.filter_by(user_id=user_id).first()
        tran = TechnicalTraining.query.filter_by(
            user_id=user_id).first()
        diploma = Diploma.query.filter_by(user_id=user_id).first()
        wb = Workbook()
        ws = wb.active
        headers = ['Nombre', 'Apellido', 'Cédula de Identidad', 'Correo Electrónico Actual', 'Dirección', 'Ciudad', 'Provincia', 'País', '¿Está completo?', 'Género', 'Edad', 'Fecha de Nacimiento', 'Tipo de Sangre', 'Donante de Sangre', 'Idioma', 'Número de Teléfono', 'Provincia', 'Tipo de Discapacidad', 'Discapacidad Específica', 'Centro de Formación', 'Otro Entrenamiento Vocacional', 'Entrenamiento','Entrenamiento Adicional', 'Centro de Estudiantes', 'Licenciatura o Técnico 1', 'Licenciatura o Técnico 2', 'Maestría 1', 'Maestría 2', 'Doctorado' , 'Instituto o Centro de Formación Técnica', 'Educación y Entrenamiento Profesional', 'Entrenamiento Vocacional o Adicional', 'Título', 'Entrenamiento Técnico', 'Centros de Formación Vocacional']

        ws.append(headers)
        row = [user.first_name, user.last_name, user.id_card, user.current_email_address, user.address, user.city, user.province, user.country, str(user.is_filled), info.gender, str(info.edad), info.date_of_birth, info.blood_type, info.blood_donor, info.language, info.home_number, info.province]
        row.append(dis.type or 'Ninguna')
        row.append(dis.specific_disability or 'Ninguna')
        row.append(inst.trainingcenter or 'Ninguno')
        row.append(inst.othervocationaltraining or 'Ninguno')
        row.append(inst.training or 'Ninguno')
        row.append(inst.addtraining or 'Ninguno')
        row.append(uni.student_center or 'Ninguno')
        row.append(uni.bachelor_or_technician_1 or 'Ninguno')
        row.append(uni.bachelor_or_technician_2 or 'Ninguno')
        row.append(uni.mastery_1 or 'Ninguno')
        row.append(uni.mastery_2 or 'Ninguno')
        row.append(uni.doctrate or 'Ninguno')
        row.append(uni.institute_or_technical_training_center or 'Ninguno')
        row.append(uni.professional_education_and_training or 'Ninguno')
        row.append(uni.vocational_training_or_additional_training or 'Ninguno')
        row.append(deg.degree or 'Ninguno')
        row.append(tran.technicalTraining or 'Ninguno')
        row.append(voc.Vocationaltrainingcenters or 'Ninguno')
        ws.append(row)

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, mode='w') as zip_file:
            zip_file.writestr(
                f'{user.first_name}_Information.xlsx', excel_buffer.getvalue())
            for file_name, file_data in [
                (f'{user.first_name}_diploma_File_' +
                diploma.filename_diploma_image, diploma.diploma_image),
                (f'{user.first_name}_identity_proof_' +
                diploma.filename_identity_proof, diploma.identity_proof),
                (f'{user.first_name}_personal_photo_' +
                diploma.filename_personal_photo, diploma.personal_photo)
            ]:
                zip_file.writestr(file_name, file_data)

        zip_buffer.seek(0)

        response = make_response(zip_buffer.getvalue())
        response.headers.set('Content-Type', 'application/zip')
        response.headers.set('Content-Disposition', 'attachment',
                            filename=f'UserRecord_{user.id_card}_Excel_Record.zip')
        return response

from flask import make_response
from openpyxl import Workbook

@admin_router.route('/diplomas/download/EXCEL', methods=['POST', 'GET'])
@login_required
def download_zip_excel_globally():
    users = db.session.query(User, OtherInformation, Disability,Institution,University,DegreeProgram,TechnicalTraining,Vocationaltrainingcenters,Diploma).join(OtherInformation).join(Disability).join(Institution).join(University).join(DegreeProgram).join(TechnicalTraining).join(Vocationaltrainingcenters).join(Diploma).all()

    wb = Workbook()
    ws = wb.active

    headers = ['Nombre', 'Apellido', 'Cédula de Identidad', 'Correo Electrónico Actual', 'Dirección', 'Ciudad', 'Provincia', 'País', '¿Está completo?', 'Género', 'Edad', 'Fecha de Nacimiento', 'Tipo de Sangre', 'Donante de Sangre', 'Idioma', 'Número de Teléfono', 'Provincia', 'Tipo de Discapacidad', 'Discapacidad Específica', 'Centro de Formación', 'Otro Entrenamiento Vocacional', 'Entrenamiento','Entrenamiento Adicional', 'Centro de Estudiantes', 'Licenciatura o Técnico 1', 'Licenciatura o Técnico 2', 'Maestría 1', 'Maestría 2', 'Doctorado' , 'Instituto o Centro de Formación Técnica', 'Educación y Entrenamiento Profesional', 'Entrenamiento Vocacional o Adicional', 'Título', 'Entrenamiento Técnico', 'Centros de Formación Vocacional', 'Descargar archivo']
    ws.append(headers)
    for user, info, dis, inst, uni, deg, tran, voc ,dip in users:
        row = [user.first_name, user.last_name, user.id_card, user.current_email_address, user.address, user.city, user.province, user.country, str(user.is_filled), info.gender, str(info.edad), info.date_of_birth, info.blood_type, info.blood_donor, info.language, info.home_number, info.province]
        row.append(dis.type or 'Ninguna')
        row.append(dis.specific_disability or 'Ninguna')
        row.append(inst.trainingcenter or 'Ninguno')
        row.append(inst.othervocationaltraining or 'Ninguno')
        row.append(inst.training or 'Ninguno')
        row.append(inst.addtraining or 'Ninguno')
        row.append(uni.student_center or 'Ninguno')
        row.append(uni.bachelor_or_technician_1 or 'Ninguno')
        row.append(uni.bachelor_or_technician_2 or 'Ninguno')
        row.append(uni.mastery_1 or 'Ninguno')
        row.append(uni.mastery_2 or 'Ninguno')
        row.append(uni.doctrate or 'Ninguno')
        row.append(uni.institute_or_technical_training_center or 'Ninguno')
        row.append(uni.professional_education_and_training or 'Ninguno')
        row.append(uni.vocational_training_or_additional_training or 'Ninguno')
        row.append(deg.degree or 'Ninguno')
        row.append(tran.technicalTraining or 'Ninguno')
        row.append(voc.Vocationaltrainingcenters or 'Ninguno')
        technical_training_link = dip.adddownload or 'Ninguno'
        if technical_training_link != 'Ninguno':
            technical_training_link = f'=HYPERLINK("http://127.0.0.1:5000{technical_training_link}","https://127.0.0.1:5000{technical_training_link}")'
        row.append(technical_training_link)
        ws.append(row)
    

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    response = make_response(excel_buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=registro_global_español.xlsx'
    return response


@admin_router.route('/FormControl/', methods=['POST', 'GET'])
@login_required
def form1():
    if request.method=='POST':
        onoff = request.form['onoff']
        form = Form.query.first()
        if onoff == "on":
            form.is_confirmed=True
            form.message="Submission is on"
            db.session.commit()
            return redirect(url_for('admin_model.admin_dashboard'))
        else: 
            form.is_confirmed=False
            form.message="Submission is off"
            db.session.commit()
            return redirect(url_for('admin_model.admin_dashboard'))
    return redirect(url_for('admin_model.admin_dashboard'))
